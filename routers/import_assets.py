"""
routers/import_assets.py — Import d'assets depuis un fichier Excel
"""

import io
import json
from datetime import datetime
from typing import Optional, Literal

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from database import get_connection

router = APIRouter(prefix="/api/import", tags=["Import"])

# ── Colonnes du template ──────────────────────────────────────────────
COLUMNS = [
    ("client_nom",          "Client *",               True,  "Nom du client (créé automatiquement si absent)"),
    ("site_nom",            "Site *",                 True,  "Nom du site (créé automatiquement si absent)"),
    ("nom_interne",         "Nom interne *",          True,  "Nom d'inventaire de l'asset (ex: SRV-METZ-01)"),
    ("type_equipement",     "Type d'équipement *",    True,  "Sélectionner dans la liste"),
    ("niveau_criticite",    "Criticité *",            True,  "faible / moyen / eleve / critique"),
    ("fabricant",           "Fabricant *",            True,  "Doit exister dans le référentiel — erreur bloquante sinon"),
    ("modele",              "Modèle",                 False, "Texte libre — doit correspondre exactement au nom dans le référentiel Modèles"),
    ("os_nom",              "OS (nom)",               False, "Sélectionner dans la liste — ex: Windows Server"),
    ("os_version",          "OS (version normalisée)",False, "Sélectionner dans la liste — ex: 2022"),
    ("version_os_libre",    "Version OS libre",       False, "Si version absente du référentiel — corrélation informatif"),
    ("firmware",            "Firmware",               False, "Sélectionner dans la liste si applicable"),
    ("statut_operationnel", "Statut opérationnel",    False, "actif / inactif / maintenance / hors_service"),
    ("adresse_ip",          "Adresse IP",             False, "ex: 192.168.1.10"),
    ("adresse_mac",         "Adresse MAC",            False, "ex: AA:BB:CC:DD:EE:FF"),
    ("hostname",            "Hostname",               False, "Nom DNS ou NetBIOS"),
    ("numero_serie",        "Numéro de série",        False, ""),
    ("date_installation",   "Date installation",      False, "Format: YYYY-MM-DD"),
    ("date_fin_garantie",   "Fin de garantie",        False, "Format: YYYY-MM-DD"),
    ("notes",               "Notes",                  False, "Informations libres"),
]

EXEMPLE_ROWS = [
    ["Administration Pénitentiaire", "MA-Metz", "SRV-DC01", "Serveur", "eleve",
     "Microsoft", "Windows Server 2022", "Windows Server", "2022", "", "",
     "actif", "192.168.1.10", "", "srv-dc01.local", "SN-12345",
     "2023-01-15", "2026-01-15", "Contrôleur de domaine principal"],
    ["Administration Pénitentiaire", "MA-Metz", "NAS-METZ-01", "NAS", "moyen",
     "Synology", "DiskStation Manager", "DSM (DiskStation Manager)", "7.2.2",
     "7.2.2-72806 Update 3", "", "actif", "192.168.1.20", "", "", "SN-67890",
     "", "", "NAS de sauvegarde"],
    ["Administration Pénitentiaire", "MA-Metz", "PC-GREFFE-01", "PC", "moyen",
     "Microsoft", "Windows 11", "Windows 11", "24H2", "", "",
     "actif", "192.168.1.30", "AA:BB:CC:DD:EE:01", "pc-greffe-01", "",
     "2024-06-01", "", "PC du greffe"],
    ["Administration Pénitentiaire", "MA-Nancy", "FW-NANCY-01", "Pare-feu", "critique",
     "Fortinet", "FortiGate", "", "", "", "FortiOS 7.4.3",
     "actif", "10.0.0.1", "", "fw-nancy", "FG-SN-001",
     "2022-03-10", "2025-03-10", "Pare-feu périmétrique"],
    ["DataVault", "Siège Paris", "CAM-HALL-01", "Caméra Axis", "faible",
     "Axis", "", "", "", "", "AXIS OS 11.11.7",
     "actif", "192.168.2.50", "00:40:8C:AA:BB:CC", "", "",
     "", "", "Caméra hall d'entrée"],
]

# ── Styles ────────────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", start_color="1F3864")
REQUIRED_FILL  = PatternFill("solid", start_color="2E4A7A")
OPTIONAL_FILL  = PatternFill("solid", start_color="2D2D2D")
EXEMPLE_FILL   = PatternFill("solid", start_color="1A3A2A")
HEADER_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT      = Font(name="Arial", size=9)
EXEMPLE_FONT   = Font(name="Arial", size=9, color="90EE90", italic=True)
THIN_BORDER    = Border(
    left=Side(style="thin", color="444444"),
    right=Side(style="thin", color="444444"),
    bottom=Side(style="thin", color="666666"),
)


def _load_referentiels(conn) -> dict:
    """Charge tous les référentiels nécessaires depuis la BDD."""
    ref = {}
    with conn.cursor() as cur:
        cur.execute("SELECT id, nom FROM product_vendors ORDER BY nom")
        ref["vendors"] = cur.fetchall()

        cur.execute("SELECT id, nom, vendor_id FROM product_models ORDER BY nom")
        ref["models"] = cur.fetchall()

        cur.execute("""
            SELECT DISTINCT os_nom FROM os_versions
            WHERE type_produit = 'os' ORDER BY os_nom
        """)
        ref["os_noms"] = [r["os_nom"] for r in cur.fetchall()]

        cur.execute("""
            SELECT DISTINCT version FROM os_versions
            WHERE type_produit = 'os' AND version IS NOT NULL ORDER BY os_nom, version
        """)
        ref["os_versions"] = [r["version"] for r in cur.fetchall() if r["version"]]

        cur.execute("""
            SELECT id, os_nom, version, nvd_product FROM os_versions
            WHERE type_produit = 'os' ORDER BY os_nom, version
        """)
        ref["os_versions_full"] = cur.fetchall()

        cur.execute("""
            SELECT id, CONCAT(os_nom, IFNULL(CONCAT(' ', version), '')) as label
            FROM os_versions WHERE type_produit = 'firmware' ORDER BY os_nom
        """)
        ref["firmwares"] = cur.fetchall()

        cur.execute("SELECT id, label, code FROM equipment_types ORDER BY label")
        ref["equipment_types"] = cur.fetchall()

    return ref


def _build_workbook(ref: dict, avec_exemple: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Feuille principale ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Assets"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A2"

    # En-têtes
    for col_idx, (key, label, required, tooltip) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if tooltip:
            from openpyxl.comments import Comment
            comment = Comment(tooltip, "Import")
            comment.width = 200
            comment.height = 60
            cell.comment = comment

    ws.row_dimensions[1].height = 35

    # ── Feuilles de référence (masquées) ─────────────────────────────
    def _make_ref_sheet(name: str, values: list, col_label: str) -> str:
        ws_ref = wb.create_sheet(name)
        ws_ref.sheet_state = "hidden"
        ws_ref.cell(row=1, column=1, value=col_label)
        for i, v in enumerate(values, 2):
            ws_ref.cell(row=i, column=1, value=str(v))
        return f"'{name}'!$A$2:$A${max(len(values)+1, 2)}"

    vendors_range    = _make_ref_sheet("_Fabricants",  [v["nom"] for v in ref["vendors"]], "Fabricant")
    # Modèles : trop nombreux pour une liste déroulante — texte libre validé côté serveur
    os_noms_range    = _make_ref_sheet("_OS_Noms",     ref["os_noms"], "OS Nom")
    os_vers_range    = _make_ref_sheet("_OS_Versions", ref["os_versions"], "OS Version")
    fw_range         = _make_ref_sheet("_Firmwares",   [f["label"] for f in ref["firmwares"]], "Firmware")
    types_range      = _make_ref_sheet("_Types",       [t["label"] for t in ref["equipment_types"]], "Type")
    crit_range       = _make_ref_sheet("_Criticite",   ["faible", "moyen", "eleve", "critique"], "Criticité")
    statut_range     = _make_ref_sheet("_Statuts",     ["actif", "inactif", "maintenance", "hors_service"], "Statut")

    # ── Data validations ──────────────────────────────────────────────
    MAX_ROWS = 5002

    def _add_dv(col_letter: str, source: str, prompt_title: str, prompt_body: str):
        dv = DataValidation(
            type="list", formula1=source, allow_blank=True,
            showInputMessage=True, showErrorMessage=True,
        )
        dv.prompt = prompt_body
        dv.promptTitle = prompt_title
        dv.error = "Valeur non autorisée. Utilisez la liste déroulante."
        dv.errorTitle = "Valeur invalide"
        dv.sqref = f"{col_letter}2:{col_letter}{MAX_ROWS}"
        ws.add_data_validation(dv)

    col_map = {key: get_column_letter(i+1) for i, (key, *_) in enumerate(COLUMNS)}

    _add_dv(col_map["type_equipement"],    types_range,   "Type d'équipement", "Sélectionner un type")
    _add_dv(col_map["niveau_criticite"],   crit_range,    "Criticité",          "faible/moyen/eleve/critique")
    _add_dv(col_map["fabricant"],          vendors_range, "Fabricant",          "Sélectionner un fabricant")

    _add_dv(col_map["os_nom"],             os_noms_range, "OS",                 "Sélectionner un OS")
    _add_dv(col_map["os_version"],         os_vers_range, "Version OS",         "Sélectionner une version")
    _add_dv(col_map["firmware"],           fw_range,      "Firmware",           "Sélectionner un firmware")
    _add_dv(col_map["statut_operationnel"],statut_range,  "Statut",             "actif/inactif/maintenance/hors_service")

    # ── Largeurs colonnes ─────────────────────────────────────────────
    widths = {
        "client_nom": 25, "site_nom": 20, "nom_interne": 20,
        "type_equipement": 18, "niveau_criticite": 12, "fabricant": 18,
        "modele": 20, "os_nom": 25, "os_version": 20, "version_os_libre": 25,
        "firmware": 22, "statut_operationnel": 16, "adresse_ip": 15,
        "adresse_mac": 18, "hostname": 20, "numero_serie": 18,
        "date_installation": 15, "date_fin_garantie": 15, "notes": 30,
    }
    for key, width in widths.items():
        ws.column_dimensions[col_map[key]].width = width

    # ── Lignes d'exemple ──────────────────────────────────────────────
    if avec_exemple:
        for row_idx, row_data in enumerate(EXEMPLE_ROWS, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = EXEMPLE_FONT
                cell.fill = EXEMPLE_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row_idx].height = 18

    # ── Onglet coloré ─────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "1F3864"

    return wb


def _xlsx_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/template")
def get_template():
    """Génère le template Excel vierge avec listes déroulantes."""
    conn = get_connection()
    try:
        ref = _load_referentiels(conn)
    finally:
        conn.close()
    wb = _build_workbook(ref, avec_exemple=False)
    return _xlsx_response(wb, "template_import.xlsx")


@router.get("/template-exemple")
def get_template_exemple():
    """Génère le template Excel avec 5 lignes d'exemple."""
    conn = get_connection()
    try:
        ref = _load_referentiels(conn)
    finally:
        conn.close()
    wb = _build_workbook(ref, avec_exemple=True)
    return _xlsx_response(wb, "template_exemple.xlsx")


# ── Preview ───────────────────────────────────────────────────────────

def _parse_date(val) -> Optional[str]:
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
    except Exception:
        return None


@router.post("/preview")
async def preview_import(file: UploadFile = File(...)):
    """
    Valide le fichier Excel sans insérer en base.
    Retourne un rapport détaillé.
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Fichier Excel invalide ou corrompu.")

    if "Assets" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Feuille 'Assets' introuvable dans le fichier.")

    ws = wb["Assets"]
    conn = get_connection()
    try:
        ref = _load_referentiels(conn)
    finally:
        conn.close()

    vendor_names  = {v["nom"].strip().lower(): v for v in ref["vendors"]}
    model_names   = {m["nom"].strip().lower(): m for m in ref["models"]}
    type_labels   = {t["label"].strip().lower(): t for t in ref["equipment_types"]}
    os_noms_set   = {n.lower() for n in ref["os_noms"]}
    os_vers_set   = {v.lower() for v in ref["os_versions"]}
    fw_labels     = {f["label"].strip().lower(): f for f in ref["firmwares"]}

    col_keys = [key for key, *_ in COLUMNS]
    rows_data = list(ws.iter_rows(min_row=2, values_only=True))

    errors   = []
    warnings = []
    valid    = 0
    new_clients = set()
    new_sites   = set()
    duplicates  = []

    conn2 = get_connection()
    try:
        with conn2.cursor() as cur:
            cur.execute("SELECT nom FROM clients")
            existing_clients = {r["nom"].strip().lower() for r in cur.fetchall()}
            cur.execute("SELECT s.nom, c.nom as client_nom FROM sites s JOIN clients c ON c.id = s.client_id")
            existing_sites = {(r["client_nom"].strip().lower(), r["nom"].strip().lower()) for r in cur.fetchall()}
            cur.execute("""
                SELECT a.nom_interne, s.nom as site_nom, c.nom as client_nom
                FROM assets a
                JOIN sites s ON s.id = a.site_id
                JOIN clients c ON c.id = s.client_id
            """)
            existing_assets = {
                (r["client_nom"].strip().lower(), r["site_nom"].strip().lower(), r["nom_interne"].strip().lower())
                for r in cur.fetchall()
            }
    finally:
        conn2.close()

    for row_idx, row in enumerate(rows_data, 2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_dict = {col_keys[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(min(len(col_keys), len(row)))}
        row_errors   = []
        row_warnings = []

        # Champs obligatoires
        for key in ["client_nom", "site_nom", "nom_interne", "type_equipement", "niveau_criticite", "fabricant"]:
            if not row_dict.get(key):
                label = next(l for k, l, *_ in COLUMNS if k == key)
                row_errors.append(f"Champ obligatoire manquant : {label}")

        if row_errors:
            errors.append({"ligne": row_idx, "erreurs": row_errors})
            continue

        # Fabricant
        fab = row_dict.get("fabricant", "").lower()
        if fab and fab not in vendor_names:
            row_errors.append(f"Fabricant '{row_dict['fabricant']}' introuvable dans le référentiel.")

        # Type équipement
        eq_type = row_dict.get("type_equipement", "").lower()
        if eq_type and eq_type not in type_labels:
            row_errors.append(f"Type d'équipement '{row_dict['type_equipement']}' introuvable.")

        # Criticité
        if row_dict.get("niveau_criticite") not in ["faible", "moyen", "eleve", "critique"]:
            row_errors.append(f"Criticité '{row_dict['niveau_criticite']}' invalide.")

        # OS nom
        os_nom = row_dict.get("os_nom", "").lower()
        if os_nom and os_nom not in os_noms_set:
            row_warnings.append(f"OS '{row_dict['os_nom']}' non trouvé dans le référentiel — sera ignoré.")

        # OS version libre
        if row_dict.get("version_os_libre") and not row_dict.get("os_nom"):
            row_warnings.append("Version OS libre sans OS nom — le préfixe OS ne sera pas ajouté.")
        if row_dict.get("version_os_libre"):
            row_warnings.append("Version OS libre → corrélation en mode informatif.")

        # Firmware
        fw = row_dict.get("firmware", "").lower()
        if fw and fw not in fw_labels:
            row_warnings.append(f"Firmware '{row_dict['firmware']}' non trouvé — sera ignoré.")

        # Nouveaux clients/sites
        client_key = row_dict["client_nom"].lower()
        site_key   = (client_key, row_dict["site_nom"].lower())
        if client_key not in existing_clients:
            new_clients.add(row_dict["client_nom"])
        if site_key not in existing_sites:
            new_sites.add(f"{row_dict['client_nom']} → {row_dict['site_nom']}")

        # Doublon
        asset_key = (client_key, row_dict["site_nom"].lower(), row_dict["nom_interne"].lower())
        if asset_key in existing_assets:
            duplicates.append({"ligne": row_idx, "asset": row_dict["nom_interne"], "site": row_dict["site_nom"]})

        if row_errors:
            errors.append({"ligne": row_idx, "erreurs": row_errors})
        else:
            valid += 1
            if row_warnings:
                warnings.append({"ligne": row_idx, "avertissements": row_warnings})

    return {
        "valides":      valid,
        "erreurs":      len(errors),
        "avertissements": len(warnings),
        "bloquant":     len(errors) > 0,
        "detail_erreurs":      errors,
        "detail_avertissements": warnings,
        "nouveaux_clients": sorted(new_clients),
        "nouveaux_sites":   sorted(new_sites),
        "doublons":         duplicates,
    }


# ── Confirm ───────────────────────────────────────────────────────────

@router.post("/confirm")
async def confirm_import(
    file: UploadFile = File(...),
    on_duplicate: str = Form("skip"),
):
    """
    Importe le fichier Excel en base.
    on_duplicate: 'skip' ou 'update'
    """
    if on_duplicate not in ("skip", "update"):
        raise HTTPException(status_code=400, detail="on_duplicate doit être 'skip' ou 'update'")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Fichier Excel invalide.")

    ws = wb["Assets"]
    col_keys = [key for key, *_ in COLUMNS]
    rows_data = list(ws.iter_rows(min_row=2, values_only=True))

    conn = get_connection()
    try:
        ref = _load_referentiels(conn)
        vendor_map = {v["nom"].strip().lower(): v["id"] for v in ref["vendors"]}
        model_map  = {m["nom"].strip().lower(): m["id"] for m in ref["models"]}
        type_map   = {t["label"].strip().lower(): t["id"] for t in ref["equipment_types"]}
        fw_map     = {f["label"].strip().lower(): f["id"] for f in ref["firmwares"]}

        # Résoudre os_version_id depuis (os_nom, version)
        os_lookup = {}
        for ov in ref["os_versions_full"]:
            key = (ov["os_nom"].lower(), (ov["version"] or "").lower())
            os_lookup[key] = ov["id"]

        stats = {"inseres": 0, "mis_a_jour": 0, "ignores": 0, "erreurs": 0}
        logs  = []

        with conn.cursor() as cur:
            # Cache clients/sites créés dans ce run
            client_cache = {}
            site_cache   = {}

            # Charger clients/sites existants
            cur.execute("SELECT id, nom FROM clients")
            for r in cur.fetchall():
                client_cache[r["nom"].strip().lower()] = r["id"]

            cur.execute("SELECT s.id, s.nom, s.client_id FROM sites s")
            for r in cur.fetchall():
                site_cache[(r["client_id"], r["nom"].strip().lower())] = r["id"]

            for row_idx, row in enumerate(rows_data, 2):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue

                row_dict = {
                    col_keys[i]: (str(row[i]).strip() if row[i] is not None else "")
                    for i in range(min(len(col_keys), len(row)))
                }

                # Champs obligatoires
                missing = [key for key in ["client_nom", "site_nom", "nom_interne", "fabricant"] if not row_dict.get(key)]
                if missing:
                    stats["erreurs"] += 1
                    logs.append({"ligne": row_idx, "statut": "erreur", "message": f"Champs manquants: {missing}"})
                    continue

                # Fabricant
                fab_id = vendor_map.get(row_dict["fabricant"].lower())
                if not fab_id:
                    stats["erreurs"] += 1
                    logs.append({"ligne": row_idx, "statut": "erreur", "message": f"Fabricant '{row_dict['fabricant']}' introuvable"})
                    continue

                # Client
                client_key = row_dict["client_nom"].lower()
                if client_key not in client_cache:
                    cur.execute("INSERT INTO clients (nom) VALUES (%s)", (row_dict["client_nom"],))
                    conn.commit()
                    client_cache[client_key] = cur.lastrowid
                    logs.append({"ligne": row_idx, "statut": "info", "message": f"Client créé : {row_dict['client_nom']}"})
                client_id = client_cache[client_key]

                # Site
                site_key = (client_id, row_dict["site_nom"].lower())
                if site_key not in site_cache:
                    cur.execute("INSERT INTO sites (client_id, nom) VALUES (%s, %s)", (client_id, row_dict["site_nom"]))
                    conn.commit()
                    site_cache[site_key] = cur.lastrowid
                    logs.append({"ligne": row_idx, "statut": "info", "message": f"Site créé : {row_dict['site_nom']}"})
                site_id = site_cache[site_key]

                # Résolutions FK
                model_id     = model_map.get(row_dict.get("modele", "").lower())
                eq_type_id   = type_map.get(row_dict.get("type_equipement", "").lower())
                fw_id        = fw_map.get(row_dict.get("firmware", "").lower())

                os_version_id = None
                if row_dict.get("os_nom") and row_dict.get("os_version"):
                    os_version_id = os_lookup.get((row_dict["os_nom"].lower(), row_dict["os_version"].lower()))

                # version_os texte libre
                version_os = None
                if not os_version_id and row_dict.get("version_os_libre"):
                    prefix = row_dict.get("os_nom", "")
                    libre  = row_dict["version_os_libre"]
                    version_os = f"{prefix} {libre}".strip() if prefix else libre
                elif row_dict.get("os_nom") and not row_dict.get("os_version") and not row_dict.get("version_os_libre"):
                    version_os = row_dict["os_nom"]

                # Criticité & statut
                criticite = row_dict.get("niveau_criticite", "moyen")
                if criticite not in ("faible", "moyen", "eleve", "critique"):
                    criticite = "moyen"
                statut_op = row_dict.get("statut_operationnel", "actif")
                if statut_op not in ("actif", "inactif", "maintenance", "hors_service"):
                    statut_op = "actif"

                # Vérifier doublon
                cur.execute("""
                    SELECT id FROM assets
                    WHERE site_id = %s AND nom_interne = %s
                """, (site_id, row_dict["nom_interne"]))
                existing = cur.fetchone()

                if existing:
                    if on_duplicate == "skip":
                        stats["ignores"] += 1
                        logs.append({"ligne": row_idx, "statut": "ignoré", "message": f"Doublon ignoré : {row_dict['nom_interne']}"})
                        continue
                    else:
                        cur.execute("""
                            UPDATE assets SET
                                vendor_id = %s, model_id = %s,
                                equipment_type_id = %s, fw_version_id = %s,
                                os_version_id = %s, version_os = %s,
                                adresse_ip = %s, adresse_mac = %s,
                                hostname = %s, numero_serie = %s,
                                niveau_criticite = %s, statut_operationnel = %s,
                                date_installation = %s, date_fin_garantie = %s,
                                notes = %s
                            WHERE id = %s
                        """, (
                            fab_id, model_id or None,
                            eq_type_id or None, fw_id or None,
                            os_version_id, version_os,
                            row_dict.get("adresse_ip") or None,
                            row_dict.get("adresse_mac") or None,
                            row_dict.get("hostname") or None,
                            row_dict.get("numero_serie") or None,
                            criticite, statut_op,
                            _parse_date(row_dict.get("date_installation")),
                            _parse_date(row_dict.get("date_fin_garantie")),
                            row_dict.get("notes") or None,
                            existing["id"],
                        ))
                        conn.commit()
                        stats["mis_a_jour"] += 1
                        logs.append({"ligne": row_idx, "statut": "mis à jour", "message": f"{row_dict['nom_interne']}"})
                        continue

                # type_equipement (enum legacy) — dériver depuis label
                type_legacy = "autre"
                if eq_type_id:
                    eq_type_obj = next((t for t in ref["equipment_types"] if t["id"] == eq_type_id), None)
                    if eq_type_obj:
                        code = eq_type_obj.get("code", "autre")
                        valid_codes = {t["code"] for t in ref["equipment_types"]}
                        type_legacy = code if code in valid_codes else "autre"

                cur.execute("""
                    INSERT INTO assets (
                        site_id, vendor_id, model_id, equipment_type_id,
                        fw_version_id, os_version_id,
                        nom_interne, type_equipement,
                        adresse_ip, adresse_mac, hostname, numero_serie,
                        version_os, niveau_criticite, statut_operationnel,
                        date_installation, date_fin_garantie, notes
                    ) VALUES (
                        %s, %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s
                    )
                """, (
                    site_id, fab_id, model_id or None, eq_type_id or None,
                    fw_id or None, os_version_id,
                    row_dict["nom_interne"], type_legacy,
                    row_dict.get("adresse_ip") or None,
                    row_dict.get("adresse_mac") or None,
                    row_dict.get("hostname") or None,
                    row_dict.get("numero_serie") or None,
                    version_os, criticite, statut_op,
                    _parse_date(row_dict.get("date_installation")),
                    _parse_date(row_dict.get("date_fin_garantie")),
                    row_dict.get("notes") or None,
                ))
                conn.commit()
                stats["inseres"] += 1
                logs.append({"ligne": row_idx, "statut": "inséré", "message": f"{row_dict['nom_interne']} → {row_dict['site_nom']}"})

        return {"stats": stats, "logs": logs}

    finally:
        conn.close()